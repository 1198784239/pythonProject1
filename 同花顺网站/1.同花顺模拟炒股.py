import csv
import requests
from lxml import etree
import re
import openpyxl
import pyecharts.options as opts
from pyecharts.charts import Line



url = 'http://moni.10jqka.com.cn/zyl.shtml'
headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0',
"Cookie":'log=; v=A3Pn2vNiEQZBw90G2nwpuDtOAnyYqAWVwT5LuSUQzAbIbJ1irXiXutEM29o2'
           }
def get_html(url):
    response = requests.get(url, headers=headers)
    res = response.text
    return res
# print(res)
# 请求头
def re_html(res):
    partt = re.compile(r'<p class="list_p(.*?)"',re.S)
    ranklist = re.compile(r'class="tab_p3">(.*?)</p>',re.S)
    name = re.compile(r'target="_blank">(.*?)<',re.S)
    win = re.compile(r'<span class="(gre|red)">(.*?)</span>',re.S)
    money = re.compile(r'class="tab_p4">(.*?)</p></td>',re.S)
    look = re.compile(r'class="tab_p4">(\d|\d\d)</p>',re.S)
    rank1 = re.findall(partt,res)
    ranklast = re.findall(ranklist,res)
    name1 = re.findall(name,res)
    win1 = re.findall(win,res)
    look1 = re.findall(look,res)
    money1 = re.findall(money,res)
    rank1.extend(ranklast)
    # print(look1)
    monenList1 = []
    for win in range(0,100):#0,3,6
        winx = win*3
        monenList1.append(win1[winx][1])
    monenList=[]
    for i in range(1,101):
        if i ==1:
            m = i*5
            monenList.append(money1[m])# 5 13 21
        else:
            m +=8
            monenList.append(money1[m])
    print(len(rank1),len(name1[1:]),len(look1),len(monenList),len(monenList1))
    dd = [list(x) for x in zip(rank1,name1[1:],monenList1,monenList,look1)]
    print(dd)

# xpath解析
def xpath(res):
    LIST = []
    # 浏览器会对html文本进行一定的规范化，所以会自动在路径中加入tbody，导致读取失败，在此处直接在路径中去除tbody
    Data =etree.HTML(res)
    Rank = []
    for num_line in range(1,101):
        try:
            rank = Data.xpath(f'//*[@id="myTab_Content2"]/table[2]/tr{[num_line]}/td[1]/p/@class')
            name = Data.xpath(f'//*[@id="myTab_Content2"]/table[2]/tr{[num_line]}/td[2]/p/a/text()')
            win = Data.xpath(f'//*[@id="myTab_Content2"]/table[2]/tr{[num_line]}/td[4]/p/span/text()')
            money = Data.xpath(f'//*[@id="myTab_Content2"]/table[2]/tr{[num_line]}/td[9]/p/text()')
            look = Data.xpath(f'//*[@id="myTab_Content2"]/table[2]/tr{[num_line]}/td[10]/p/text()')
            list = [name[0], win[0], money[0], look[0]]
            LIST.append(list)
            if num_line <=10:
                rank_f = rank[0].replace('list_p','')
                list.insert(0,rank_f)
            else:
                # Rank.append(num_line)
                # print(num_line)
                list.insert(0,num_line)
        except:
            pass
    print(LIST)
    return LIST
# 保存为excel
def Save_excel(LIST):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '排行'
    ws['B1'] = '名字'
    ws['C1'] = '收益'
    ws['D1'] = '资产'
    ws['E1'] = '点赞'
    for i in range(len(LIST)):
        ws.append(LIST[i])
    wb.save('shuju.xlsx')

# 保存为csv
def save_csv(LIST):
    with open('shujuya.csv','w',encoding='utf-8') as f:
        writer = csv.writer(f)
        headers = ['排行','名字','收益','资产','点赞']
        writer.writerow(headers)
        for row in LIST:
            writer.writerow(row)
# 图表分析
def chart_html(LIST):
    x_data = []
    y_data = []
    for i in range(0,6):
        xl = LIST[i][1]
        x_data.append(xl)
        yl = LIST[i][3]
        y_data.append(yl)

    print(x_data)
    print(y_data)

    # x_data = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    # y_data = [820, 932, 901, 934, 1290, 1330, 1320]

    (
        Line()
        .add_xaxis(xaxis_data=x_data)
        .add_yaxis(
            series_name="邮件营销",
            stack="总量",
            y_axis=y_data,
            label_opts=opts.LabelOpts(is_show=False),
        )
        .add_yaxis(
            series_name="联盟广告",
            stack="总量",
            y_axis=[220, 182, 191, 234, 290, 330, 310],
            label_opts=opts.LabelOpts(is_show=False),
        )
        .add_yaxis(
            series_name="视频广告",
            stack="总量",
            y_axis=[150, 232, 201, 154, 190, 330, 410],
            label_opts=opts.LabelOpts(is_show=False),
        )
        .add_yaxis(
            series_name="直接访问",
            stack="总量",
            y_axis=[320, 332, 301, 334, 390, 330, 320],
            label_opts=opts.LabelOpts(is_show=False),
        )
        .add_yaxis(
            series_name="搜索引擎",
            stack="总量",
            y_axis=[820, 932, 901, 934, 1290, 1330, 1320],
            label_opts=opts.LabelOpts(is_show=False),
        )
        .set_global_opts(
            title_opts=opts.TitleOpts(title="折线图堆叠"),
            tooltip_opts=opts.TooltipOpts(trigger="axis"),
            yaxis_opts=opts.AxisOpts(
                type_="value",
                axistick_opts=opts.AxisTickOpts(is_show=True),
                splitline_opts=opts.SplitLineOpts(is_show=True),
            ),
            xaxis_opts=opts.AxisOpts(type_="category", boundary_gap=False),
        )
        .render("stacked_line_chart.html")
    )


if __name__ == '__main__':
    get_html(url)
    xpath(get_html(url))
    chart_html(xpath(get_html(url)))
    # re_html(get_html(url))
    # Save_excel(xpath(get_html(url)))
    # save_csv(xpath(get_html(url)))