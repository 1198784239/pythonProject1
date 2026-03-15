import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
import re

# ===================== 核心配置（仅需确认此处） =====================
root_folder = r"C:\Users\李克聪\Desktop\特命"  # 原始数据文件夹路径
main_output_folder = "研发总表"  # 最终结果主文件夹名
company_folders = ["特敏"]

# 列名配置（与你的Excel保持一致）
work_hour_col = "工时（小时）"
date_col = "日期（年-月-日）"
PROJECT_NAME_COL = "项目名称"
year_col = "年月"  # 表格中存储年份的列名
work_content_col = "工时内容"  # 明确工时内容列名

# ===================== 样式定义 =====================
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
header_font = Font(bold=True, size=11)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ===================== 列宽配置（强制设置，优先级最高）=====================
# 键：列名（支持模糊匹配），值：目标宽度
COL_WIDTH_SETTINGS = {
    "年月": 12,
    "姓名": 18,
    "项目代码": 18,
    "项目名称": 50,
    "日期": 18,  # 模糊匹配"日期（年-月-日）"
    "工时（小时）": 25,  # 强制设置工时列宽
    "工时内容": 60,  # 强制设置工时内容列宽
}


# ===================== 工具函数 =====================
def safe_file_name(name):
    """去除Windows非法字符，生成安全的文件名/文件夹名"""
    illegal_chars = [":", "/", "\\", "?", "*", "\"", "<", ">", "|"]
    for char in illegal_chars:
        name = name.replace(char, "")
    return name.strip().replace(" ", "_")


def extract_year_from_str(year_str):
    """从字符串中提取4位年份（如202301→2023、2024-05→2024、2025年→2025）"""
    year_match = re.search(r'(\d{4})', str(year_str))
    if year_match:
        return year_match.group(1)
    return "未知年份"


def get_target_col_width(col_name):
    """根据列名模糊匹配目标宽度"""
    for key, width in COL_WIDTH_SETTINGS.items():
        if key in col_name:
            return width
    return 20  # 默认宽度


# ===================== 核心处理函数 =====================
def process_company_excel(company_name, excel_path):
    """处理单个公司的Excel文件，自动识别年份创建子文件夹"""
    print(f"\n🔄 正在处理：{company_name} - {excel_path}")

    # 1. 读取源数据
    df = pd.read_excel(
        excel_path,
        dtype=str,
        usecols=None,
        keep_default_na=False
    )

    # 打印实际列名，方便排查匹配问题
    print(f"📋 表格实际列名：{df.columns.tolist()}")

    # 校验关键列
    required_cols = [year_col, "姓名", PROJECT_NAME_COL, work_hour_col]
    for col in required_cols:
        if col not in df.columns:
            print(f"❌ {company_name}：缺失列 '{col}'，跳过处理")
            return

    # 格式处理
    df[year_col] = df[year_col].astype(str).str.strip()
    df["姓名"] = df["姓名"].astype(str).str.strip()
    if date_col in df.columns:
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')

    # 2. 创建公司级文件夹
    main_output_path = os.path.join(os.path.dirname(root_folder), main_output_folder)
    company_output_path = os.path.join(main_output_path, company_name)
    Path(main_output_path).mkdir(exist_ok=True)
    Path(company_output_path).mkdir(exist_ok=True)

    # 3. 按「年月+姓名」分组拆分
    group_keys = [year_col, "姓名"]
    for (year_month, name), group_df in df.groupby(group_keys):
        # 排序
        if date_col in group_df.columns:
            group_df = group_df.sort_values(by=date_col, ascending=True)
            group_df[date_col] = group_df[date_col].dt.strftime("%Y-%m-%d").fillna("")

        # 重置索引
        group_df = group_df.reset_index(drop=True)
        actual_cols = group_df.columns.tolist()
        col_count = len(actual_cols)

        if len(group_df) == 0:
            print(f"⚠️  跳过空数据：{name}_{year_month}")
            continue

        # 构建最终表格
        title_row = pd.DataFrame([["研发项目工时记录表"] + [""] * (col_count - 1)], columns=actual_cols)
        header_row = pd.DataFrame([actual_cols], columns=actual_cols)
        total_row = pd.DataFrame([[""] * col_count], columns=actual_cols)
        total_row.iloc[0, 0] = "合计："

        # 计算工时合计
        if work_hour_col in actual_cols:
            group_df[work_hour_col] = pd.to_numeric(group_df[work_hour_col], errors='coerce').fillna(0)
            total_hours = group_df[work_hour_col].sum()
            total_row.iloc[0, actual_cols.index(work_hour_col)] = total_hours

        final_df = pd.concat([title_row, header_row, group_df, total_row], ignore_index=True)

        # 生成安全的文件名并保存
        year_month_safe = safe_file_name(year_month)
        folder_year = extract_year_from_str(year_month)
        year_output_path = os.path.join(company_output_path, folder_year)
        Path(year_output_path).mkdir(exist_ok=True)

        file_name = f"{name}_{year_month_safe}.xlsx"
        file_path = os.path.join(year_output_path, file_name)
        final_df.to_excel(file_path, index=False, header=False)

        # 4. 美化Excel格式（重点修复列宽）
        wb = load_workbook(file_path)
        ws = wb.active
        total_row_num = ws.max_row

        # 合并标题行
        last_col = get_column_letter(col_count)
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'].font = Font(bold=True, size=15)
        ws['A1'].alignment = center_align
        ws.row_dimensions[1].height = 20

        # 表头行样式
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # 数据行样式
        for row in range(3, total_row_num):
            for col_idx in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align

        # 合计行样式
        ws.merge_cells(f'A{total_row_num}:E{total_row_num}')
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=total_row_num, column=col_idx)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

        # ========== 强制设置列宽（核心修复） ==========
        print(f"🔧 开始设置列宽...")
        for col_idx, col_name in enumerate(actual_cols, 1):
            col_letter = get_column_letter(col_idx)
            # 获取目标宽度（模糊匹配）
            target_width = get_target_col_width(col_name)
            # 强制设置列宽，关闭自动适配
            ws.column_dimensions[col_letter].width = target_width
            ws.column_dimensions[col_letter].auto_size = False
            print(f"   ✅ 列 '{col_name}' → 宽度 {target_width}")

        # 行高设置
        for row in range(2, total_row_num + 1):
            ws.row_dimensions[row].height = 22

        # 保存并关闭
        wb.save(file_path)
        wb.close()
        print(f"✅ 生成成功：{file_path}")


# ===================== 遍历所有公司文件夹（主程序） =====================
for company in company_folders:
    company_path = os.path.join(root_folder, company)
    if not os.path.isdir(company_path):
        print(f"⚠️  原始数据文件夹不存在：{company_path}，跳过")
        continue

    # 查找所有Excel文件
    excel_files = list(Path(company_path).glob("*.xlsx")) + list(Path(company_path).glob("*.xls"))
    if not excel_files:
        print(f"⚠️  {company} 文件夹下未找到Excel文件，跳过")
        continue

    # 处理每个Excel文件
    for excel_file in excel_files:
        process_company_excel(company, str(excel_file))

print(f"\n🎉 全部处理完成！最终结果总文件夹：{os.path.join(os.path.dirname(root_folder), main_output_folder)}")